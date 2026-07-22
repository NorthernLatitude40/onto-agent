import openpyxl
import os
import json
import datetime
import opencc
from openpyxl.styles import PatternFill, Alignment
from core.tools.tools import validate_design_json

# 初始化转换器：繁体转简体
converter = opencc.OpenCC("t2s")
# 定义灰色背景样式
gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

def normalize_key(key):
    """将键名统一转换为简体中文，去除前后空格"""
    if key is None: return ""
    return converter.convert(str(key).strip())

def run_test_with_file(file_path="raw.json"):
    print(f"--- 測試開始：讀取檔案 {file_path} ---")
    current_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(current_dir, file_path)

    if not os.path.exists(file_path):
        print(f"錯誤：找不到檔案 {file_path}")
        return

    with open(file_path, "r", encoding="utf-8") as f:
        raw_content = f.read()

    # 調用工具進行校驗與序列化
    result = validate_design_json.func(raw_content)
    print(f"\n驗證結果輸出:\n{result}")

    generate_excel_from_json(result)

def generate_excel_from_json(json_str, template_name="template_1.xlsx"):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(current_dir, template_name)

    if not os.path.exists(template_path):
        print(f"错误：未找到模板文件 '{template_name}'")
        return
        
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(current_dir, f"详细设计书_生成版_{timestamp}.xlsx")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # 解析 JSON 字串為字典
    full_json = json.loads(json_str)
    
    # 獲取 items 列表
    items = full_json.get("items", [])

    # 自動定位表頭行
    header_row = None
    target_headers = ["no", "項目名称"]
    for row in range(1, 15):
        row_values = [normalize_key(cell.value) for cell in ws[row] if cell.value is not None]
        if any(h in row_values for h in target_headers):
            header_row = row
            break

    if not header_row:
        print("错误：未能识别到表头行。")
        return

    header_map = {normalize_key(cell.value): cell.column for cell in ws[header_row] if cell.value}

    # 填入數據
    start_row = header_row + 1
    for i, item in enumerate(items): # 修正：這裡遍歷的是 list
        row = start_row + i

        if item.get("is_group"):
            ws.cell(row=row, column=1, value=item.get("item_name"))
            max_col = ws.max_column
            for merged_range in list(ws.merged_cells.ranges):
                if merged_range.min_row <= row <= merged_range.max_row:
                    ws.unmerge_cells(str(merged_range))
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.fill = gray_fill
        else:
            # 將 JSON 的 key 映射到 Excel
            # 你的 JSON key 為英文，需對應 Excel 的中文標題
            mapping = {
                "no": "No",
                "item_name": "項目名称",
                "category": "分類",
                "required": "必須",
                "field_code": "桁数",
                "format": "フォーマット",
                "table": "テーブル",
                "field_name": "フィールド",
                "remarks": "備考"
            }
            
            for key, col_idx in header_map.items():
                # 找到對應的 JSON key
                json_key = next((k for k, v in mapping.items() if normalize_key(v) == key), None)
                if json_key and json_key in item:
                    ws.cell(row=row, column=col_idx, value=item[json_key])

    wb.save(output_path)
    print(f"成功！文件已生成至: {output_path}")

if __name__ == "__main__":
    run_test_with_file()