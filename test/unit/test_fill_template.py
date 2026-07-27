import datetime
import json
import os

import opencc
import openpyxl
from openpyxl.styles import Alignment, PatternFill

# 初始化转换器：繁体转简体
converter = opencc.OpenCC("t2s")
# 定义灰色背景样式
gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")


def normalize_key(key):
    """将键名统一转换为简体中文，去除前后空格"""
    return converter.convert(str(key).strip())


def load_data_from_json(filename="excel_data.json"):
    """从当前脚本同目录下的 JSON 文件加载数据，并提取 '画面项目'"""
    current_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(current_dir, filename)

    if not os.path.exists(file_path):
        print(f"错误：未找到数据文件 '{filename}'")
        return []

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            items_list = data.get("画面项目", []) if isinstance(data, dict) else data

            processed_data = []
            for item in items_list:
                if isinstance(item, dict):
                    processed_data.append(
                        {normalize_key(k): v for k, v in item.items()}
                    )
                else:
                    processed_data.append(item)
            return processed_data
    except (RuntimeError, OSError) as e:
        print(f"错误：解析 JSON 文件失败: {e}")
        return []


def fill_template_dynamically(data, template_name="template_1.xlsx"):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(current_dir, template_name)

    if not os.path.exists(template_path):
        print(f"错误：未找到模板文件 '{template_name}'")
        return

    timestamp = datetime.datetime.now(datetime.UTC).strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(current_dir, f"详细设计书_生成版_{timestamp}.xlsx")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # --- 1. 新增：读取 JSON 中的头部信息并填入 ---
    # 假设你已经把整个 JSON 加载为 full_data
    # 需要在外部调用时把 full_data 传入，或者在这里重新读取一次
    with open(
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "excel_data.json"),
        "r",
        encoding="utf-8",
    ) as f:
        full_json = json.load(f)
        header_info = full_json.get("系统头部信息", {})

    # 遍历所有单元格，如果单元格的值匹配“逻辑名称”、“系统名称”等，就填入对应的值
    # 这样无论这些字在 Excel 哪个位置，都能被自动定位填入
    for row in range(1, 10):  # 头部信息通常在上面几行
        for col in range(1, ws.max_column + 1):
            cell_val = normalize_key(ws.cell(row=row, column=col).value)
            if cell_val in header_info:
                ws.cell(row=row, column=col + 1, value=header_info[cell_val])

    # 自动定位表头行
    header_row = None
    target_headers = ["no", "項目名称"]
    for row in range(1, 15):
        row_values = [
            normalize_key(cell.value) for cell in ws[row] if cell.value is not None
        ]
        if any(h in row_values for h in target_headers):
            header_row = row
            break

    if not header_row:
        print("错误：未能识别到表头行。")
        return

    # 构建映射字典
    header_map = {}
    for cell in ws[header_row]:
        if cell.value:
            header_map[normalize_key(cell.value)] = cell.column

    # 填入数据
    start_row = header_row + 1
    for i, item in enumerate(data):
        row = start_row + i

        # 判断是否为机构行
        if item.get("is_group"):
            # 1. 填入机构名称到第一列
            ws.cell(row=row, column=1, value=item.get("項目名称"))
            
            # 2. 获取当前实际最大列数
            max_col = ws.max_column
            
            # 3. 核心：删除该行所有可能冲突的现有合并单元格
            # 我们遍历所有已合并区域，如果该区域与当前行重叠，则将其移除
            for merged_range in list(ws.merged_cells.ranges):
                if merged_range.min_row <= row <= merged_range.max_row:
                    ws.unmerge_cells(str(merged_range))
            
            # 4. 执行我们需要的整行合并
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
            
            # 5. 设置样式
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.fill = gray_fill
            
            continue # 确保机构行处理完后跳过普通逻辑
        else:
            # 普通行填入数据
            for key, value in item.items():
                if key == "is_group":
                    continue
                normalized_k = normalize_key(key)
                if normalized_k in header_map and value is not None:
                    ws.cell(row=row, column=header_map[normalized_k], value=value)

    wb.save(output_path)
    print(f"成功！文件已生成至: {output_path}")


if __name__ == "__main__":
    raw_data = load_data_from_json("excel_data.json")
    flattened_data = []

    for entry in raw_data:
        if "机构" in entry:
            # 插入机构行
            flattened_data.append({"項目名称": entry.get("机构"), "is_group": True})
            # 展开项目
            if "項目" in entry and isinstance(entry["項目"], list):
                for item in entry["項目"]:
                    norm_item = {normalize_key(k): v for k, v in item.items()}
                    norm_item["is_group"] = False
                    flattened_data.append(norm_item)
        else:
            entry["is_group"] = False
            flattened_data.append(entry)

    if flattened_data:
        fill_template_dynamically(flattened_data)
