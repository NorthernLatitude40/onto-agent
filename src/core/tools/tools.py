import datetime
import json
import os
from pathlib import Path

import httpx
import opencc
import openpyxl
from langchain_core.tools import tool
from openpyxl.styles import Alignment, PatternFill

from src.config.config import settings

# 引入 Pydantic 结构
from src.ingestion.schema.screen_item import DesignDocument
from src.ontology.screen_dict import HeaderSemanticResolver, SheetSemanticResolver


@tool
def search_official_knowledge_base(query: str) -> str:
    """RAG 查询官方售票知识库"""
    url = f"{settings.ANYTHINGLLM_BASE_URL}/workspace/{settings.WORKSPACE_SLUG}/chat"
    payload = {
        "message": query,
        "mode": "query",
        "model": "current",
        "temperature": 0.0,
    }
    headers = {
        "Authorization": f"Bearer {settings.ANYTHINGLLM_API_KEY}",
        "Content-Type": "application/json",
    }

    try:
        r = httpx.post(url, json=payload, headers=headers, timeout=15)
        r.raise_for_status()
        data = r.json()

        sources = data.get("sources", [])
        if sources:
            return "\n\n".join(
                src.get("text", "").strip() for src in sources if src.get("text")
            )

        return data.get("textResponse", "")

    except (ImportError, Exception) as e:
        return f"RAG error: {e}"


@tool
def get_weather(city: str) -> str:
    """获取指定城市的实时天气信息"""
    if "东京" in city or "tokyo" in city.lower():
        return "东京：晴 18°C"
    if "台北" in city or "taipei" in city.lower():
        return "台北：阴 22°C"
    return f"{city} 天气未知"


@tool
def validate_design_json(raw_json_str: str) -> str:
  """接收 LLM 生成的原始 JSON 字串，並強制轉化為標準的設計書 JSON。

  如果格式不符合 DesignDocument 規範，則會拋出錯誤。
  """
  try:
    # 1. 處理可能的 Markdown 代碼塊格式
    clean_json = raw_json_str.replace("```json", "").replace("```", "").strip()

    # 2. Pydantic 核心校驗：將野生 JSON 轉為物件並強制校驗
    data = safe_json_loads(clean_json)
    design_doc = DesignDocument(**data)

    # 3. 輸出標準化後的 JSON (使用 by_alias=True 確保輸出時全是當初定義的中文/日文/No 欄位名稱)
    return json.dumps(
        design_doc.model_dump(by_alias=True), ensure_ascii=False, indent=2
    )
  except (ImportError, Exception) as e:
    raise ValueError(f"设计书 JSON 校验失败：{e}")


# 初始化转换器：繁体转简体
converter = opencc.OpenCC("t2s")
# 定义灰色背景样式
gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")


def normalize_key(key):
    """去除键名前后空格，不进行繁简体转换"""
    if key is None:
        return ""
    return str(key).strip()

# resolver 只需在模組載入時建立一次，之後重複使用
_header_resolver = HeaderSemanticResolver()
_sheet_resolver = SheetSemanticResolver()

def generate_excel(json_str: str, template_name: str = "template_1.xlsx") -> str:
    """
    根据设计书 JSON 生成 Excel 文件。並提供下載鏈接。
    """
    path = _build_excel(json_str, template_name)
    filename = os.path.basename(path)
    base_url = "http://127.0.0.1:8000"
    download_url = f"{base_url}/files/{filename}"

    return (
        f"Excel 已生成！請點擊下方連結下載：\n[📥 點擊下載詳細設計書]({download_url})"
    )


def _build_excel(
    json_str: str,
    template_name: str = "template_1.xlsx",
) -> str:
    current_dir = Path(__file__).resolve().parent.parent
    export_dir = current_dir.parent.parent / "exports"
    export_dir.mkdir(exist_ok=True)
    template_path = current_dir.parent / "ingestion" / "templates" / template_name

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"模板不存在：{template_path}")

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"詳細設計書_生成版_{timestamp}.xlsx"
    output_path = export_dir / filename

    wb = openpyxl.load_workbook(template_path)

    # 用別名機制找出目標分頁，取代原本寫死的 `target_sheet_name = "画面項目"`
    ws = _sheet_resolver.get_sheet(wb, default_index=0)

    try:
        full_json = safe_json_loads(json_str)
    except json.JSONDecodeError as e:
        raise ValueError(f"JSON 格式錯誤：{e}")

    # 相容兩種來源：畫面設計書（screen_item schema）用「畫面項目」，
    # Parser → 設計書 bridge（design_doc_builder）用 "items"
    items = full_json.get("畫面項目") or full_json.get("items", [])

    # 自動定位表頭列 + 建立 canonical column map
    scan_rows = [
        [cell.value for cell in row]
        for row in ws.iter_rows(min_row=1, max_row=10)
    ]
    header_row_idx, col_map = _header_resolver.find_header_row(scan_rows, min_matches=3)
    if header_row_idx is None:
        raise ValueError("Excel 模板中未找到可辨識的表頭。")
    header_row = header_row_idx + 1  # ws.iter_rows 是 0-based，openpyxl 儲存格是 1-based

    # 定義群組行背景色 (灰色)
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # 填入數據
    start_row = header_row + 1
    for i, item in enumerate(items):
        row = start_row + i

        # 防呆層：把 item 的 key 統一 resolve 成 canonical key。
        # design_doc_builder.py 產出的 item 已經是 canonical key，這裡是 no-op；
        # 若來源是舊格式（中日文 key），這裡仍會被正確轉換。
        normalized_item = {
            (_header_resolver.resolve(k) or k): v for k, v in item.items()
        }

        if normalized_item.get("is_group"):
            ws.cell(row=row, column=1, value=normalized_item.get("item_name"))
            max_col = ws.max_column
            for merged_range in list(ws.merged_cells.ranges):
                if merged_range.min_row <= row <= merged_range.max_row:
                    ws.unmerge_cells(str(merged_range))
            ws.merge_cells(
                start_row=row, start_column=1, end_row=row, end_column=max_col
            )
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.fill = gray_fill
        else:
            for canonical_key, col_idx in col_map.items():
                if canonical_key in normalized_item:
                    ws.cell(
                        row=row,
                        column=col_idx + 1,
                        value=normalized_item[canonical_key],
                    )

    wb.save(output_path)
    return str(output_path)

def safe_json_loads(data):
    """確保無論傳入的是 str 還是 dict，都能回傳 dict"""
    if isinstance(data, dict):
        return data
    elif isinstance(data, (str, bytes, bytearray)):
        return json.loads(data)
    else:
        raise TypeError(f"無法解析的 JSON 型態: {type(data)}")