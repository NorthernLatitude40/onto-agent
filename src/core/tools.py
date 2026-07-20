import httpx
from langchain_core.tools import tool
from src.config.config import ANYTHINGLLM_BASE_URL, ANYTHINGLLM_API_KEY, WORKSPACE_SLUG

# 引入 Pydantic 结构
from src.ingestion.demo.design_schema import DesignDocument
import json
import openpyxl
import os
import json
import datetime
import opencc
from openpyxl.styles import PatternFill, Alignment
from pathlib import Path

# 將 JSON 的 key 映射到 Excel
# 你的 JSON key 為英文，需對應 Excel 的中文標題
# 1. 修改映射表為 {Excel表頭: JSON鍵名}
mapping = {
    "No": "No",
    "项目名称": "项目名称",
    "分类": "分类",
    "必须": "必须",
    "栏目号码": "栏目号码",
    "格式": "格式",
    "表格": "表格",
    "栏域": "栏域",
    "備考": "備考",
}


@tool
def search_official_knowledge_base(query: str) -> str:
    """RAG 查询官方售票知识库"""
    url = f"{ANYTHINGLLM_BASE_URL}/workspace/{WORKSPACE_SLUG}/chat"
    payload = {
        "message": query,
        "mode": "query",
        "model": "current",
        "temperature": 0.0,
    }
    headers = {
        "Authorization": f"Bearer {ANYTHINGLLM_API_KEY}",
        "Content-Type": "application/json",
    }

    try:
        r = httpx.post(url, json=payload, headers=headers, timeout=15)
        r.raise_for_status()
        data = r.json()

        sources = data.get("sources", [])
        if sources:
            return "\n\n".join(
                src.get("text", "").strip() for src in sources if src.get("text")
            )

        return data.get("textResponse", "")

    except Exception as e:
        return f"RAG error: {e}"


@tool
def get_weather(city: str) -> str:
    """获取指定城市的实时天气信息"""
    if "东京" in city or "tokyo" in city.lower():
        return "东京：晴 18°C"
    if "台北" in city or "taipei" in city.lower():
        return "台北：阴 22°C"
    return f"{city} 天气未知"


@tool
def validate_design_json(raw_json_str: str) -> str:
    """
    接收 LLM 生成的原始 JSON 字串，並強制轉化為標準的設計書 JSON。
    如果格式不符合 DesignDocument 規範，則會拋出錯誤。
    """
    try:
        # 1. 處理可能的 Markdown 代碼塊格式
        clean_json = raw_json_str.replace("```json", "").replace("```", "").strip()

        # 2. Pydantic 核心校驗：將野生 JSON 轉為物件並強制校驗
        data = json.loads(clean_json)
        design_doc = DesignDocument(**data)

        # 3. 輸出標準化後的 JSON
        return json.dumps(
            design_doc.model_dump(),
            ensure_ascii=False,
            indent=2,
        )
    except Exception as e:
        raise ValueError(f"设计书 JSON 校验失败：{e}")


# 初始化转换器：繁体转简体
converter = opencc.OpenCC("t2s")
# 定义灰色背景样式
gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")


def normalize_key(key):
    """将键名统一转换为简体中文，去除前后空格"""
    if key is None:
        return ""
    return converter.convert(str(key).strip())


@tool
def generate_excel(json_str: str, template_name: str = "template_1.xlsx") -> str:
    """
    根据设计书 JSON 生成 Excel 文件。並提供下載鏈接。
    """
    path = _build_excel(json_str, template_name)
    filename = os.path.basename(path)
    base_url = "http://0.0.0.0:8000"
    download_url = f"{base_url}/files/{filename}"

    return (
        f"Excel 已生成！請點擊下方連結下載：\n[📥 點擊下載詳細設計書]({download_url})"
    )


def _build_excel(
    json_str: str,
    template_name: str = "template_1.xlsx",
) -> str:
    current_dir = Path(__file__).resolve().parent
    export_dir = current_dir.parent / "exports"
    export_dir.mkdir(exist_ok=True)
    template_path = current_dir.parent / "ingestion" / "demo" / template_name

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"模板不存在：{template_path}")

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"详细设计书_生成版_{timestamp}.xlsx"
    output_path = export_dir / filename

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    try:
        full_json = json.loads(json_str)
    except json.JSONDecodeError as e:
        raise ValueError(f"JSON 格式错误：{e}")

    # 獲取 items 列表
    items = full_json.get("畫面項目", [])

    # 自動定位表頭行
    header_row = None
    target_headers = ["no", "项目名称"]
    for row in range(1, 15):
        row_values = [
            normalize_key(cell.value) for cell in ws[row] if cell.value is not None
        ]
        if any(h in row_values for h in target_headers):
            header_row = row
            break

    if not header_row:
        raise ValueError("Excel 模板中未找到表头。")

    header_map = {
        normalize_key(cell.value): cell.column for cell in ws[header_row] if cell.value
    }

    # 填入數據
    start_row = header_row + 1
    for i, item in enumerate(items):  # 修正：這裡遍歷的是 list
        row = start_row + i

        if item.get("is_group"):
            ws.cell(row=row, column=1, value=item.get("item_name"))
            max_col = ws.max_column
            for merged_range in list(ws.merged_cells.ranges):
                if merged_range.min_row <= row <= merged_range.max_row:
                    ws.unmerge_cells(str(merged_range))
            ws.merge_cells(
                start_row=row, start_column=1, end_row=row, end_column=max_col
            )
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.fill = gray_fill
        else:

            # 直接根據 header_map 找到對應的欄位
            for header_name, col_idx in header_map.items():
                # 根據 Excel 表頭從 mapping 取得對應的 JSON Key
                # 注意：這裡的 header_name 是你在 Excel 讀取到的原始值
                json_key = mapping.get(header_name)

                if json_key and json_key in item:
                    ws.cell(row=row, column=col_idx, value=item[json_key])

    wb.save(output_path)
    return str(output_path)
